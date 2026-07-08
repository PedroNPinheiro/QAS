import io
import re
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Type

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from sqlalchemy import Boolean as SABoolean, Integer as SAInteger, cast, func, or_, select
from sqlalchemy.orm import Session

from .audit import diff_changes, log_audit
from .auth import get_current_user, require_admin
from .permissions import editable_fields, require_create, require_view
from .config import settings
from .database import get_db
from .models import Attachment, NotificationRecipient, User
from .schemas import Page
from .sequences import next_reference


def _serialize(obj):
    creator = getattr(obj, "created_by", None)
    obj.created_by_name = creator.full_name if creator else None
    return obj


# Export header labels that plain title-casing gets wrong
HEADER_OVERRIDES = {
    "po": "PO",
    "ler_code": "LER Code",
    "egar": "e-GAR",
    "root_cause": "Root Cause Analysis",
    "cost": "Cost (€)",
    "invoiced_value": "Invoiced Value (€)",
    "quantity_kg": "Quantity (kg)",
    "act_notified": "Communicated to ACT",
    "insurance_notified": "Insurance Participated",
    "derogation": "Product Derogation",
    "first_derogation_po": "First Derogation PO",
    "last_derogation_po": "Last Derogation PO",
    "date_detected": "Date",
    "created_by_name": "Created By",
    "nature": "Nature of Injury",
    "return_note": "Return Note Nº",
}

# Human labels for enum values, applied only to enum-backed fields
ENUM_FIELDS = {"status", "severity", "risk_level"}
ENUM_LABELS = {
    "open": "Open",
    "in_progress": "In progress",
    "closed": "Closed",
    "on_time": "On time",
    "delayed": "Delayed",
    "concluded": "Concluded",
    "minor": "Minor",
    "major": "Major",
    "critical": "Critical",
    "first_aid": "First aid",
    "serious": "Serious",
    "fatal": "Fatal",
    "low": "Low",
    "medium": "Medium",
    "high": "High",
}


def _header_label(field: str) -> str:
    return HEADER_OVERRIDES.get(field, field.replace("_", " ").title())


def _cell_value(field: str, value):
    if value is None:
        return None
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, datetime):
        # Excel cannot store timezone-aware datetimes
        return value.astimezone().replace(tzinfo=None) if value.tzinfo else value
    if field in ENUM_FIELDS:
        return ENUM_LABELS.get(str(value), str(value))
    return value


def _build_xlsx(title: str, fields: list[str], rows: list[dict]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1C5CAB")
    widths: list[int] = []
    for col, field in enumerate(fields, start=1):
        label = _header_label(field)
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
        widths.append(len(label))

    for r, row in enumerate(rows, start=2):
        for col, field in enumerate(fields, start=1):
            value = _cell_value(field, row.get(field))
            ws.cell(row=r, column=col, value=value)
            if value is not None:
                if isinstance(value, datetime):
                    length = 16
                elif isinstance(value, date):
                    length = 10
                else:
                    length = min(len(str(value)), 55)
                widths[col - 1] = max(widths[col - 1], length)

    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width + 3
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(fields))}{max(len(rows) + 1, 1)}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def create_crud_router(
    *,
    model,
    ref_prefix: str,
    entity_type: str,
    create_schema: Type[BaseModel],
    update_schema: Type[BaseModel],
    read_schema: Type[BaseModel],
    search_fields: tuple[str, ...],
    date_field: str,
    ref_style: str = "monthly",
    display_name: str = "record",
    frontend_path: str = "",
    filter_fields: tuple[str, ...] = (),
    notify: str | None = None,  # None | "fixed" (recipients table) | "choose" (creator picks users)
) -> APIRouter:
    """Build a standard CRUD router: paginated list with search/filter/sort,
    Excel export, get, create (with auto reference), update, delete (admin)."""

    router = APIRouter()
    sortable = {c.name for c in model.__table__.columns}

    def ref_matches_period(reference: str, d: date) -> bool:
        """Does this reference belong to the date's month/year sequence?
        Unrecognised formats are left alone (treated as matching)."""
        if ref_style == "yearly":
            m = re.match(r"^(\d+)_(\d{4})$", reference)
            return not m or int(m.group(2)) == d.year
        m = re.match(rf"^{ref_prefix}(\d{{2}})(\d{{2}})\.\d+$", reference)
        return not m or (int(m.group(1)) == d.month and int(m.group(2)) == d.year % 100)

    date_col = getattr(model, date_field)

    def apply_filters(stmt, q: str | None, status: str | None, request: Request | None = None):
        if q:
            clauses = [getattr(model, f).ilike(f"%{q}%") for f in search_fields]
            clauses.append(model.reference.ilike(f"%{q}%"))
            stmt = stmt.where(or_(*clauses))
        if status and hasattr(model, "status"):
            stmt = stmt.where(model.status == status)
        if request is not None:
            for key, value in request.query_params.items():
                if not key.startswith("f_") or not value:
                    continue
                fname = key[2:]
                if fname not in filter_fields:
                    continue
                col = getattr(model, fname)
                if isinstance(model.__table__.columns[fname].type, SABoolean):
                    stmt = stmt.where(col.is_(value.lower() == "true"))
                else:
                    stmt = stmt.where(col == value)
            date_from = request.query_params.get("date_from")
            date_to = request.query_params.get("date_to")
            try:
                if date_from:
                    stmt = stmt.where(date_col >= date.fromisoformat(date_from))
                if date_to:
                    stmt = stmt.where(date_col < date.fromisoformat(date_to) + timedelta(days=1))
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid date filter")
        return stmt

    def apply_sort(stmt, sort: str):
        if not sort and ref_style == "yearly":
            # yearly registers (quality tests): the number IS the order
            year_num = cast(func.split_part(model.reference, "_", 2), SAInteger)
            seq_num = cast(func.split_part(model.reference, "_", 1), SAInteger)
            return stmt.order_by(year_num.desc(), seq_num.desc())
        sort = sort or f"-{date_field}"
        field = sort.lstrip("-")
        if field not in sortable:
            field, sort = date_field, f"-{date_field}"
        col = getattr(model, field)
        # id as tiebreak keeps same-day records in register (import) order
        if sort.startswith("-"):
            return stmt.order_by(col.desc(), model.id.desc())
        return stmt.order_by(col.asc(), model.id.asc())

    @router.get("", response_model=Page[read_schema])
    def list_records(
        request: Request,
        q: str | None = None,
        status: str | None = None,
        page: int = Query(1, ge=1),
        page_size: int = Query(20, ge=1, le=200),
        sort: str = "",
        db: Session = Depends(get_db),
        user: User = Depends(get_current_user),
    ):
        require_view(user, entity_type)
        stmt = apply_filters(select(model), q, status, request)
        total = db.scalar(select(func.count()).select_from(stmt.subquery()))
        stmt = apply_sort(stmt, sort).offset((page - 1) * page_size).limit(page_size)
        items = [_serialize(o) for o in db.scalars(stmt).all()]
        return Page(items=items, total=total, page=page, page_size=page_size)

    @router.get("/export.xlsx")
    def export_xlsx(
        request: Request,
        q: str | None = None,
        status: str | None = None,
        sort: str = "",
        db: Session = Depends(get_db),
        user: User = Depends(get_current_user),
    ):
        require_view(user, entity_type)
        stmt = apply_sort(apply_filters(select(model), q, status, request), sort)
        rows = [
            read_schema.model_validate(_serialize(o)).model_dump()
            for o in db.scalars(stmt).all()
        ]
        # Reference first, data fields in schema order, audit columns last
        meta = {"id", "reference", "created_at", "updated_at", "created_by_name"}
        fields = (
            ["reference"]
            + [f for f in read_schema.model_fields if f not in meta]
            + ["created_at", "created_by_name"]
        )
        title = display_name if display_name != "record" else entity_type.replace("_", " ").title()
        buf = _build_xlsx(title, fields, rows)
        filename = f"{entity_type}-{date.today().isoformat()}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    @router.get("/filter-options")
    def filter_options(
        db: Session = Depends(get_db),
        user: User = Depends(get_current_user),
    ):
        require_view(user, entity_type)
        out: dict[str, list] = {}
        for fname in filter_fields:
            if isinstance(model.__table__.columns[fname].type, SABoolean):
                continue
            col = getattr(model, fname)
            values = db.scalars(
                select(col).where(col.isnot(None)).distinct().order_by(col).limit(300)
            ).all()
            out[fname] = [v for v in values if str(v).strip()]
        return out

    @router.get("/{record_id}", response_model=read_schema)
    def get_record(
        record_id: int,
        db: Session = Depends(get_db),
        user: User = Depends(get_current_user),
    ):
        require_view(user, entity_type)
        obj = db.get(model, record_id)
        if obj is None:
            raise HTTPException(status_code=404, detail="Record not found")
        return _serialize(obj)

    @router.post("", response_model=read_schema, status_code=201)
    def create_record(
        payload: create_schema,
        background: BackgroundTasks,
        db: Session = Depends(get_db),
        user: User = Depends(get_current_user),
    ):
        require_create(user, entity_type)
        data = payload.model_dump()
        notify_emails = data.pop("notify_emails", [])
        obj = model(**data)
        # Reference month follows the record's own date, not today
        obj.reference = next_reference(
            db, ref_prefix, getattr(obj, date_field, None), ref_style, model
        )
        obj.created_by_id = user.id
        db.add(obj)
        db.flush()
        db.refresh(obj)
        log_audit(
            db,
            entity_type=entity_type,
            entity_id=obj.id,
            reference=obj.reference,
            action="create",
            user=user,
        )

        # Email notifications (sent after the response, failures only logged):
        # fixed recipients for this module + the global 'all' list + any
        # addresses the creator picked on the form.
        emails = list(
            db.scalars(
                select(NotificationRecipient.email).where(
                    NotificationRecipient.entity_type.in_([entity_type, "all"])
                )
            )
        )
        if notify == "choose" and notify_emails:
            emails += [e for e in notify_emails]
        emails = list(dict.fromkeys(e.lower() for e in emails))
        db.commit()  # commit before responding so an immediate refetch sees the record

        if emails:
            from .mailer import record_email, send_email  # lazy: avoids import cycle

            meta = {"id", "created_at", "updated_at", "reference"}
            fields = [f for f in read_schema.model_fields if f not in meta]
            record = read_schema.model_validate(_serialize(obj)).model_dump()
            subject, body = record_email(
                display_name=display_name,
                record=record,
                fields=fields,
                record_path=f"{frontend_path}/{obj.id}",
            )
            background.add_task(send_email, emails, subject, body)

        return _serialize(obj)

    @router.put("/{record_id}", response_model=read_schema)
    def update_record(
        record_id: int,
        payload: update_schema,
        db: Session = Depends(get_db),
        user: User = Depends(get_current_user),
    ):
        require_view(user, entity_type)
        allowed = editable_fields(user, entity_type)
        obj = db.get(model, record_id)
        if obj is None:
            raise HTTPException(status_code=404, detail="Record not found")
        new_values = payload.model_dump(exclude_unset=True)
        if allowed is not None:
            new_values = {k: v for k, v in new_values.items() if k in allowed}
        changes = diff_changes(obj, new_values)
        for key, value in new_values.items():
            setattr(obj, key, value)
        # A date edit that moves the record into another month/year re-homes
        # the reference into that period's sequence; the old number is freed.
        if date_field in changes:
            d = getattr(obj, date_field)
            if isinstance(d, datetime):
                d = d.date()
            if d is not None and not ref_matches_period(obj.reference, d):
                new_ref = next_reference(db, ref_prefix, d, ref_style, model)
                changes["reference"] = {"from": obj.reference, "to": new_ref}
                obj.reference = new_ref
        if changes:
            log_audit(
                db,
                entity_type=entity_type,
                entity_id=obj.id,
                reference=obj.reference,
                action="update",
                user=user,
                changes=changes,
            )
        db.commit()
        db.refresh(obj)
        return _serialize(obj)

    @router.delete("/{record_id}", status_code=204)
    def delete_record(
        record_id: int,
        db: Session = Depends(get_db),
        user: User = Depends(require_admin),
    ):
        obj = db.get(model, record_id)
        if obj is None:
            raise HTTPException(status_code=404, detail="Record not found")
        attachments = db.scalars(
            select(Attachment).where(
                Attachment.entity_type == entity_type, Attachment.entity_id == record_id
            )
        ).all()
        upload_dir = Path(settings.upload_dir)
        for att in attachments:
            (upload_dir / att.stored_name).unlink(missing_ok=True)
            db.delete(att)
        log_audit(
            db,
            entity_type=entity_type,
            entity_id=obj.id,
            reference=obj.reference,
            action="delete",
            user=user,
        )
        db.delete(obj)
        db.commit()

    return router
