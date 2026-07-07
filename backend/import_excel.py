"""One-off importer for the department's historical Excel registers.

Usage:
    python import_excel.py /opt/QAS/docs            # dry-run: parse + report only
    python import_excel.py /opt/QAS/docs --apply    # write to the database
    python import_excel.py /opt/QAS/docs --apply --wipe
        # first clear ALL records/attachments/audit/counters (users are kept)

Cleaning rules (reported per row when applied):
- Dates: Excel dates used as-is; text dates parsed dd/mm/yyyy with typo fixes
  ("15/002/2023", "19/01/20026", ranges "a ..." take the first date).
- Internal/External NC without an explicit detection date use the
  communication date, else the reference's month (day 1).
- Duplicate references: identical rows are skipped; different records with a
  duplicated number get the next free number in their month (reported).
- Status: records closed in the register (implementation/closure date) or from
  years before 2026 import as closed; 2026 rows with actions in progress
  import as in_progress, otherwise open.
- LTI rows without a real date (old 2019-era summary lines) are skipped.
"""
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime, time, timezone
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete, select

from app import models
from app.database import Base, SessionLocal, engine

ISSUES: list[str] = []
PT_MONTHS = {
    "janeiro": 1, "fevereiro": 2, "março": 3, "marco": 3, "abril": 4, "maio": 5,
    "junho": 6, "julho": 7, "agosto": 8, "setembro": 9, "outubro": 10,
    "novembro": 11, "dezembro": 12,
}


def issue(msg: str) -> None:
    ISSUES.append(msg)


def txt(v, limit: int | None = None) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if not s or s.lower() == "none":
        return None
    return s[:limit] if limit else s


def num(v) -> float | None:
    if isinstance(v, (int, float)):
        return float(v)
    s = txt(v)
    if not s:
        return None
    s = s.replace("€", "").replace(",", ".").strip()
    try:
        return float(s)
    except ValueError:
        return None


def parse_date(v, context: str = "") -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = txt(v)
    if not s:
        return None
    s = s.split(" a ")[0].split(" e ")[0].strip()  # ranges: keep the first day
    s = re.sub(r"/0+(\d\d)/", r"/\1/", s)  # 15/002/2023 -> 15/02/2023
    s = re.sub(r"/(2)0+(\d{3})$", r"/2\2", s)  # 19/01/20026 -> 19/01/2026
    s = re.sub(r"^(\d{1,2})/(\d{2})(\d{4})$", r"\1/\2/\3", s)  # 19/072023 -> 19/07/2023
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    issue(f"unparseable date {v!r} ({context})")
    return None


def yes(v) -> bool:
    s = txt(v)
    if not s:
        return False
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode().lower()
    return s.startswith(("sim", "y", "x", "1", "true"))


def marker(v) -> bool:
    return txt(v) is not None


def created_ts(d: date) -> datetime:
    return datetime.combine(d, time(12, 0), tzinfo=timezone.utc)


class RefBook:
    """Tracks used references and hands out the next free number per month."""

    def __init__(self, existing: set[str]):
        self.used = set(existing)

    def claim(self, ref: str) -> bool:
        if ref in self.used:
            return False
        self.used.add(ref)
        return True

    def next_free(self, prefix: str, mm: int, yy: int) -> str:
        n = 1
        while f"{prefix}{mm:02d}{yy:02d}.{n:02d}" in self.used:
            n += 1
        ref = f"{prefix}{mm:02d}{yy:02d}.{n:02d}"
        self.used.add(ref)
        return ref


def clean_ref(raw: str, prefix: str) -> tuple[str | None, str | None]:
    """Extract a reference like NCE0226.08 from messy register cells
    ("NCE0226.08\nLote 1, Palete 1", "NCE1124,14"); returns (ref, leftover)."""
    m = re.search(rf"{prefix}\s*(\d{{4}})\s*[.,]\s*(\d{{1,3}})", raw)
    if not m:
        return None, None
    leftover = (raw[: m.start()] + " " + raw[m.end():]).replace("\n", " ").strip(" ·-—,;") or None
    return f"{prefix}{m.group(1)}.{int(m.group(2)):02d}", leftover


def ref_month(ref: str) -> tuple[int, int] | None:
    m = re.match(r"^[A-Z]+(\d{2})(\d{2})\.\d+$", ref)
    return (int(m.group(1)), int(m.group(2))) if m else None


# ------------------------------------------------------------- internal NCs

def load_internal(docs: Path, refs: RefBook) -> list[models.InternalNonConformity]:
    wb = load_workbook(docs / "Internal NC.xlsx", data_only=True)
    sheets = {
        "Internal NC (PT)": "v2023",
        "Internal NC (PT) (2024)": "v2024",
        "Internal NC (PT) (2025)": "v2024",
        "Internal NC (PT) (2026)": "v2024",
    }
    out, seen_rows = [], {}
    for sheet, version in sheets.items():
        ws = wb[sheet]
        for r in ws.iter_rows(min_row=5):
            raw_ref = txt(r[0].value)
            if not raw_ref or not raw_ref.startswith("NCI"):
                continue
            ref, ref_extra = clean_ref(raw_ref, "NCI")
            if not ref:
                issue(f"internal {raw_ref!r}: unrecognised reference — skipped")
                continue
            if version == "v2023":
                desc = txt(r[4].value) or txt(r[5].value)
                melhoria = txt(r[5].value) if txt(r[4].value) else None
                rec = dict(
                    po=txt(r[1].value, 100), project=txt(r[2].value, 255),
                    sector=txt(r[3].value, 100), description=desc,
                    cost=num(r[6].value),
                    communicated_date=parse_date(r[7].value, ref),
                    implementation_date=parse_date(r[8].value, ref),
                    notes=" · ".join(x for x in [txt(r[9].value),
                                     f"Improvement: {melhoria}" if melhoria else None] if x) or None,
                    corrective_action=None, preventive_action=None,
                    designer=None, cost_note=None,
                )
            else:
                rec = dict(
                    po=txt(r[1].value, 100), project=txt(r[2].value, 255),
                    sector=txt(r[3].value, 100), description=txt(r[4].value),
                    designer=txt(r[5].value, 255), cost=num(r[6].value),
                    cost_note=txt(r[7].value),
                    corrective_action=txt(r[8].value), preventive_action=txt(r[9].value),
                    communicated_date=parse_date(r[10].value, ref),
                    implementation_date=parse_date(r[11].value, ref),
                    notes=txt(r[12].value),
                )
            if not rec["description"]:
                issue(f"internal {ref}: no description — skipped")
                continue
            if ref_extra:
                rec["notes"] = " · ".join(x for x in [rec["notes"], f"Ref. note: {ref_extra}"] if x)

            key = (ref, rec["description"][:80])
            if key in seen_rows:
                issue(f"internal {ref}: identical duplicate row — skipped")
                continue
            seen_rows[key] = True

            mmyy = ref_month(ref)
            dup = not refs.claim(ref)
            if dup and not mmyy:
                issue(f"internal {ref}: malformed duplicate — skipped")
                continue
            date_detected = rec["communicated_date"] or (
                date(2000 + mmyy[1], mmyy[0], 1) if mmyy else None
            )
            if not date_detected:
                issue(f"internal {ref}: no usable date — skipped")
                continue
            if rec["implementation_date"] or date_detected.year < 2026:
                status = "closed"
            elif rec["corrective_action"] or rec["preventive_action"]:
                status = "in_progress"
            else:
                status = "open"
            out.append((dup, mmyy, models.InternalNonConformity(
                reference=ref, date_detected=date_detected, severity="minor",
                status=status, created_at=created_ts(date_detected),
                updated_at=created_ts(date_detected), **rec,
            )))
    # second pass: only true duplicates get renumbered, after all originals claimed
    records = []
    for dup, mmyy, obj in out:
        if dup:
            new_ref = refs.next_free("NCI", *mmyy)
            issue(f"internal {obj.reference}: duplicated number in register — imported as {new_ref}")
            obj.reference = new_ref
        records.append(obj)
    return records


# ------------------------------------------------------------- external NCs

def load_external(docs: Path, refs: RefBook) -> list[models.ExternalNonConformity]:
    wb = load_workbook(docs / "External NC.xlsx", data_only=True)
    ws = wb["Folha1"]
    out = []
    for r in ws.iter_rows(min_row=9):
        raw_ref = txt(r[1].value)
        if not raw_ref or not raw_ref.startswith("NCE"):
            continue
        ref, ref_extra = clean_ref(raw_ref, "NCE")
        if not ref:
            issue(f"external {raw_ref!r}: unrecognised reference — skipped")
            continue
        mmyy = ref_month(ref)
        d = parse_date(r[2].value, ref) or (date(2000 + mmyy[1], mmyy[0], 1) if mmyy else None)
        if not d:
            issue(f"external {ref}: no usable date — skipped")
            continue
        dup = not refs.claim(ref)
        if dup and not mmyy:
            issue(f"external {ref}: malformed duplicate — skipped")
            continue
        description = txt(r[9].value) or txt(r[7].value) or "(no description in register)"
        closure = parse_date(r[19].value, ref)
        if closure or d.year < 2026:
            status = "closed"
        elif txt(r[14].value) or txt(r[16].value):
            status = "in_progress"
        else:
            status = "open"
        out.append(models.ExternalNonConformity(
            reference=ref, date_detected=d,
            supplier=txt(r[3].value, 255) or "(not recorded)",
            po=txt(r[4].value, 100), delivery_doc=txt(r[5].value, 100),
            item_reference=txt(r[6].value, 100), item_designation=txt(r[7].value, 255),
            quantity=num(r[8].value), description=description,
            location=txt(r[10].value, 255), has_control_range=yes(r[11].value),
            communicated_date=parse_date(r[12].value, ref),
            supplier_response=txt(r[14].value), root_cause=txt(r[15].value),
            action_to_take=" · ".join(x for x in [txt(r[16].value),
                            f"Return note (prov.): {txt(r[13].value)}" if txt(r[13].value) else None] if x) or None,
            return_note=txt(r[17].value, 100),
            closure_responsible=txt(r[18].value, 255), closure_date=closure,
            notes=" · ".join(x for x in [txt(r[20].value),
                     f"Ref. note: {ref_extra}" if ref_extra else None] if x) or None,
            severity="minor", status=status,
            created_at=created_ts(d), updated_at=created_ts(d),
        ))
        out[-1] = (dup, mmyy, out[-1])
    records = []
    for dup, mmyy, obj in out:
        if dup:
            new_ref = refs.next_free("NCE", *mmyy)
            issue(f"external {obj.reference}: duplicated number in register — imported as {new_ref}")
            obj.reference = new_ref
        records.append(obj)
    return records


# ---------------------------------------------------------------- accidents

LTI_BODY = [(19, "Arm/Shoulder"), (21, "Hand"), (23, "Lumbar"), (25, "Fingers"),
            (27, "Head"), (29, "Leg/Foot"), (31, "Abdomen")]
LTI_NATURE = [(33, "Cut"), (35, "Perforation"), (37, "Muscular"),
              (39, "Hit/Projection"), (41, "Fall"), (43, "Burn")]
LTI_DEPT = [(55, "Metal"), (57, "Painting"), (59, "Vinyl"), (61, "Glass Assembly"),
            (63, "Glass Transformation"), (65, "Assembly"), (67, "VETs Assembly"),
            (69, "Petsmart Assembly"), (71, "Plumbing"), (73, "Electrics"),
            (75, "Joinery"), (77, "Finishing and Cleaning"), (79, "Packing"),
            (81, "Warehouse"), (83, "Staff")]


def load_accidents(docs: Path, refs: RefBook) -> list[models.WorkAccident]:
    wb = load_workbook(docs / "Acidentes de Trabalho (LTI).xlsx", data_only=True)
    ws = wb["LTI (d)"]
    rows = []
    for r in ws.iter_rows(min_row=5, max_col=91):
        name = txt(r[1].value)
        if not name or name.lower().startswith("total"):
            continue
        if not isinstance(r[5].value, datetime):
            issue(f"accident '{name[:30]}': no full date (old register style) — skipped")
            continue
        d = r[5].value
        hora = txt(r[6].value)
        if hora:
            m = re.match(r"^(\d{1,2})[hH:.](\d{0,2})", hora)
            if m:
                d = d.replace(hour=int(m.group(1)), minute=int(m.group(2) or 0))
        pick = lambda pairs: next((label for idx, label in pairs if marker(r[idx].value)), None)
        days = num(r[2].value) or 0
        rows.append(dict(
            occurred=d, injured=name,
            days_lost=int(days), hours_lost=num(r[3].value),
            body_part=pick(LTI_BODY), nature=pick(LTI_NATURE), dept=pick(LTI_DEPT),
            insurance=yes(r[85].value), act=yes(r[86].value),
            inability=txt(r[87].value, 255), detail=txt(r[88].value),
            corrective=txt(r[89].value), preventive=txt(r[90].value),
        ))
    rows.sort(key=lambda x: x["occurred"])
    out = []
    for x in rows:
        ref = refs.next_free("ACC", x["occurred"].month, x["occurred"].year % 100)
        out.append(models.WorkAccident(
            reference=ref, occurred_at=x["occurred"].replace(tzinfo=timezone.utc),
            injured_person=x["injured"][:255],
            department=x["dept"],
            description=x["detail"] or "(no detail in register)",
            body_part=x["body_part"], nature=x["nature"],
            severity="minor" if x["days_lost"] > 0 else "first_aid",
            days_lost=x["days_lost"], hours_lost=x["hours_lost"],
            inability=x["inability"], insurance_notified=x["insurance"],
            act_notified=x["act"], corrective_action=x["corrective"],
            preventive_action=x["preventive"], status="closed",
            created_at=created_ts(x["occurred"].date()),
            updated_at=created_ts(x["occurred"].date()),
        ))
    return out


# --------------------------------------------------------------- near misses

NM_EVENTS = [(3, "Fall"), (4, "Shock"), (5, "Break"), (6, "Cut/Perforation"),
             (7, "Fire"), (8, "Electrical Discharge"), (9, "Fire"), (10, "Other")]
NM_LOCS = [(11, "Metal"), (12, "Painting"), (13, "Vinyl"), (14, "Glass Assembly"),
           (15, "Glass Transformation"), (16, "Assembly"), (17, "VETs Assembly"),
           (18, "Petsmart Assembly"), (19, "Plumbing"), (20, "Electrics"),
           (21, "Joinery"), (22, "Finishing and Cleaning"), (23, "Packing"),
           (24, "Warehouse"), (25, "Staff"), (26, "RPP"), (27, "External Waste Park")]
NM_STATUS = {"close": "concluded", "closed": "concluded", "concluded": "concluded",
             "ongoing": "on_time", "on time": "on_time", "delayed": "delayed"}


def load_near_misses(docs: Path, refs: RefBook) -> list[models.NearMiss]:
    wb = load_workbook(docs / "Near Miss.xlsx", data_only=True)
    ws = wb["NM"]
    rows = []
    for r in ws.iter_rows(min_row=7, max_col=34):
        desc = txt(r[1].value)
        if not desc:
            continue
        d = parse_date(r[2].value, f"near miss '{desc[:25]}'")
        if not d:
            issue(f"near miss '{desc[:30]}': no date — skipped")
            continue
        pick = lambda pairs: next((label for idx, label in pairs if marker(r[idx].value)), None)
        status_raw = (txt(r[32].value) or "").lower()
        rows.append(dict(
            d=d, desc=desc, event=pick(NM_EVENTS), loc=pick(NM_LOCS),
            corrective=txt(r[28].value), preventive=txt(r[29].value),
            owner=txt(r[31].value, 255),
            status=NM_STATUS.get(status_raw, "concluded"),
            close=parse_date(r[33].value, "nm close"),
        ))
    rows.sort(key=lambda x: x["d"])
    out = []
    for x in rows:
        ref = refs.next_free("NM", x["d"].month, x["d"].year % 100)
        out.append(models.NearMiss(
            reference=ref, occurred_date=x["d"], description=x["desc"],
            event_type=x["event"], location=x["loc"], risk_level="low",
            corrective_action=x["corrective"], preventive_action=x["preventive"],
            owner=x["owner"], preventive_close_date=x["close"], status=x["status"],
            created_at=created_ts(x["d"]), updated_at=created_ts(x["d"]),
        ))
    return out


# --------------------------------------------------------------- test reports

def load_tests(docs: Path, used_test_refs: set[str]) -> list[models.TestReport]:
    wb = load_workbook(docs / "Test report and product derogation 1.xlsx", data_only=True)
    ws = wb["Folha1"]
    out, last_date = [], None
    for r in ws.iter_rows(min_row=6, max_col=12):
        raw_ref = txt(r[1].value)
        if not raw_ref:
            continue
        m = re.match(r"^0*(\d+)[_/](\d{4})$", raw_ref)
        if not m:
            issue(f"test {raw_ref!r}: unrecognised reference — skipped")
            continue
        ref = f"{int(m.group(1))}_{m.group(2)}"
        dup = ref in used_test_refs
        used_test_refs.add(ref)
        d = parse_date(r[2].value, f"test {ref}")
        if not d:
            d = last_date or date(int(m.group(2)), 1, 1)
            issue(f"test {ref}: missing date — used {d} (previous row's date)")
        last_date = d
        desc = txt(r[5].value) or txt(r[6].value) or "(no description in register)"
        out.append(models.TestReport(
            reference=ref, test_date=d, tested_by=txt(r[4].value, 255),
            description=desc, result=txt(r[6].value) if txt(r[5].value) else None,
            products_affected=txt(r[7].value), observations=txt(r[8].value),
            derogation=yes(r[9].value),
            first_derogation_po=txt(r[10].value), last_derogation_po=txt(r[11].value),
            created_at=created_ts(d), updated_at=created_ts(d),
        ))
        out[-1] = (dup, m.group(2), out[-1])
    # true duplicates get the next free number in their year
    records = []
    for dup, year, obj in out:
        if dup:
            nums = [int(x.split("_")[0]) for x in used_test_refs if x.endswith("_" + year)]
            new_ref = f"{max(nums) + 1}_{year}"
            used_test_refs.add(new_ref)
            issue(f"test {obj.reference}: duplicated number in register — imported as {new_ref}")
            obj.reference = new_ref
        records.append(obj)
    return records


# -------------------------------------------------------------------- waste

HAZ_LER = {"150202", "150110", "80111", "080111", "80113", "080113",
           "161001", "200136", "200121", "130110", "131010"}


def load_waste(docs: Path, refs: RefBook) -> list[models.WasteRecord]:
    wb = load_workbook(docs / "Registo de Produção de Resíduos.xlsx", data_only=True)
    ws = wb["Consumos"]
    rows = []
    for r in ws.iter_rows(min_row=7, max_col=11):
        wtype = txt(r[1].value, 255)
        if not wtype:
            continue
        wtype = {"RIbs": "RIBs"}.get(wtype, wtype)
        d = parse_date(r[4].value, f"waste '{wtype[:20]}'") if not (
            isinstance(r[4].value, str) and r[4].value.strip().isdigit()) else None
        if d is None:
            # year in the date column + Portuguese month name in the month column
            year_txt, month_txt = txt(r[4].value), (txt(r[5].value) or "").lower()
            month = PT_MONTHS.get(month_txt)
            if year_txt and year_txt.isdigit() and month:
                d = date(int(year_txt), month, 1)
                issue(f"waste '{wtype[:20]}' r{r[0].row}: only month given — used {d}")
        if not d:
            issue(f"waste '{wtype[:20]}' r{r[0].row}: no usable date — skipped")
            continue
        qty = num(r[6].value)
        if not qty or qty <= 0:
            issue(f"waste '{wtype[:20]}' r{r[0].row}: missing quantity — skipped")
            continue
        ler = txt(r[2].value, 20)
        ler_key = re.sub(r"\D", "", ler or "")
        rows.append(dict(
            d=d, wtype=wtype, ler=ler, desc=txt(r[3].value, 255), qty=qty,
            haz="contamin" in wtype.lower() or ler_key in HAZ_LER,
            egar=txt(r[7].value, 100), operator=txt(r[8].value, 255),
            value=num(r[9].value), notes=txt(r[10].value),
        ))
    rows.sort(key=lambda x: x["d"])
    out = []
    for x in rows:
        ref = refs.next_free("WST", x["d"].month, x["d"].year % 100)
        out.append(models.WasteRecord(
            reference=ref, collection_date=x["d"], waste_type=x["wtype"],
            ler_code=x["ler"], waste_description=x["desc"], quantity_kg=x["qty"],
            hazardous=x["haz"], egar=x["egar"], operator=x["operator"],
            invoiced_value=x["value"], notes=x["notes"],
            created_at=created_ts(x["d"]), updated_at=created_ts(x["d"]),
        ))
    return out


# --------------------------------------------------------------------- main

def sync_sequences(db) -> None:
    db.execute(delete(models.RefSequence))
    counters: dict[tuple, int] = defaultdict(int)
    monthly = {"NCI": models.InternalNonConformity, "NCE": models.ExternalNonConformity,
               "ACC": models.WorkAccident, "NM": models.NearMiss, "WST": models.WasteRecord}
    for prefix, model in monthly.items():
        for (ref,) in db.execute(select(model.reference)):
            m = re.match(rf"^{prefix}(\d{{2}})(\d{{2}})\.(\d+)$", ref)
            if m:
                key = (prefix, 2000 + int(m.group(2)), int(m.group(1)))
                counters[key] = max(counters[key], int(m.group(3)))
    for (ref,) in db.execute(select(models.TestReport.reference)):
        m = re.match(r"^(\d+)_(\d{4})$", ref)
        if m:
            key = ("TR", int(m.group(2)), 0)
            counters[key] = max(counters[key], int(m.group(1)))
    for (prefix, year, month), last in counters.items():
        db.add(models.RefSequence(prefix=prefix, year=year, month=month, last_number=last))


def main() -> None:
    docs = Path(sys.argv[1])
    apply = "--apply" in sys.argv
    wipe = "--wipe" in sys.argv
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    try:
        if wipe and apply:
            for model in (models.AuditLog, models.Attachment, models.InternalNonConformity,
                          models.ExternalNonConformity, models.TestReport, models.WorkAccident,
                          models.NearMiss, models.WasteRecord, models.RefSequence,
                          models.NotificationRecipient):
                db.execute(delete(model))
            print("Wiped records, attachments, audit log, counters (users kept).")

        existing = {ref for model in (models.InternalNonConformity, models.ExternalNonConformity,
                                      models.WorkAccident, models.NearMiss, models.WasteRecord)
                    for (ref,) in db.execute(select(model.reference))} if not wipe else set()
        refs = RefBook(existing)
        used_tests = ({ref for (ref,) in db.execute(select(models.TestReport.reference))}
                      if not wipe else set())

        batches = {
            "Internal NCs": load_internal(docs, refs),
            "External NCs": load_external(docs, refs),
            "Quality tests": load_tests(docs, used_tests),
            "Accidents": load_accidents(docs, refs),
            "Near misses": load_near_misses(docs, refs),
            "Waste records": load_waste(docs, refs),
        }

        print("\n=== IMPORT SUMMARY ===")
        for name, records in batches.items():
            years = Counter(getattr(r, "created_at").year for r in records)
            print(f"{name:15s} {len(records):5d}   {dict(sorted(years.items()))}")
        print(f"\nIssues/fixes: {len(ISSUES)}")
        for line in ISSUES:
            print("  -", line)

        if apply:
            for records in batches.values():
                db.add_all(records)
            db.flush()
            sync_sequences(db)
            db.commit()
            print("\nAPPLIED — records written and reference counters synced.")
        else:
            print("\nDRY RUN — nothing written. Re-run with --apply to import.")
    finally:
        db.close()


if __name__ == "__main__":
    main()
